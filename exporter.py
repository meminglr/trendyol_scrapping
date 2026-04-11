"""
Veri Dışa Aktarma Modülü
=========================
Çekilen verileri CSV, JSON, Excel ve SQLite formatlarında kaydeder.
"""

import csv
import json
import logging
from pathlib import Path
from typing import Union

from models import Product, Review, SearchResult

logger = logging.getLogger(__name__)


def _ensure_dir(path: str) -> Path:
    p = Path(path)
    p.mkdir(parents=True, exist_ok=True)
    return p


# ------------------------------------------------------------------ #
#  JSON                                                                #
# ------------------------------------------------------------------ #

def save_json(data: Union[list, dict, SearchResult], filepath: str):
    """Veriyi JSON dosyasına kaydeder."""
    if isinstance(data, SearchResult):
        data = data.to_dicts()
    elif isinstance(data, list) and data and isinstance(data[0], Product):
        data = [p.to_dict() for p in data]
    elif isinstance(data, list) and data and isinstance(data[0], Review):
        data = [r.to_dict() for r in data]

    Path(filepath).parent.mkdir(parents=True, exist_ok=True)
    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    logger.info(f"JSON kaydedildi: {filepath}")


# ------------------------------------------------------------------ #
#  CSV                                                                 #
# ------------------------------------------------------------------ #

def save_csv(products: list[Product], filepath: str):
    """Ürün listesini CSV dosyasına kaydeder."""
    if not products:
        logger.warning("Kaydedilecek ürün yok.")
        return

    Path(filepath).parent.mkdir(parents=True, exist_ok=True)

    fieldnames = [
        "id", "name", "brand", "category", "price", "original_price",
        "discount_pct", "currency", "rating", "review_count", "in_stock",
        "seller", "free_shipping", "cargo_days", "url", "images",
        "badges", "has_gift",
    ]

    with open(filepath, "w", newline="", encoding="utf-8-sig") as f:  # utf-8-sig: Excel uyumlu
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        for p in products:
            row = p.to_dict()
            row["images"] = " | ".join(row.get("images", []))
            row["badges"] = ", ".join(row.get("badges", []))
            writer.writerow(row)

    logger.info(f"CSV kaydedildi: {filepath} ({len(products)} ürün)")


def save_reviews_csv(reviews: list[Review], filepath: str):
    """Yorum listesini CSV dosyasına kaydeder."""
    if not reviews:
        return

    Path(filepath).parent.mkdir(parents=True, exist_ok=True)
    fieldnames = ["id", "product_id", "author", "rating", "title", "comment",
                  "date", "likes", "dislikes", "is_verified", "variant", "seller_response"]

    with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        for r in reviews:
            writer.writerow(r.to_dict())

    logger.info(f"Yorumlar CSV kaydedildi: {filepath}")


# ------------------------------------------------------------------ #
#  Excel (openpyxl)                                                    #
# ------------------------------------------------------------------ #

def save_excel(products: list[Product], filepath: str, reviews: list[Review] = None):
    """
    Ürünleri (ve isteğe bağlı yorumları) Excel dosyasına kaydeder.
    Şart: pip install openpyxl
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        logger.error("openpyxl kurulu değil. `pip install openpyxl` çalıştırın.")
        return

    wb = openpyxl.Workbook()

    # --- Ürünler Sayfası ---
    ws = wb.active
    ws.title = "Ürünler"

    headers = [
        "ID", "Ürün Adı", "Marka", "Kategori", "Fiyat (TRY)",
        "Orijinal Fiyat", "İndirim %", "Puan", "Değerlendirme",
        "Stokta", "Satıcı", "Ücretsiz Kargo", "Badge", "URL"
    ]

    # Header stili
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF6000", end_color="FF6000", fill_type="solid")
    center = Alignment(horizontal="center")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    # Veri satırları
    for row, p in enumerate(products, 2):
        ws.cell(row=row, column=1, value=p.id)
        ws.cell(row=row, column=2, value=p.name)
        ws.cell(row=row, column=3, value=p.brand)
        ws.cell(row=row, column=4, value=p.category)
        ws.cell(row=row, column=5, value=p.price)
        ws.cell(row=row, column=6, value=p.original_price)
        ws.cell(row=row, column=7, value=p.discount_pct)
        ws.cell(row=row, column=8, value=p.rating)
        ws.cell(row=row, column=9, value=p.review_count)
        ws.cell(row=row, column=10, value="Evet" if p.in_stock else "Hayır")
        ws.cell(row=row, column=11, value=p.seller)
        ws.cell(row=row, column=12, value="Evet" if p.free_shipping else "Hayır")
        ws.cell(row=row, column=13, value=", ".join(p.badges))
        ws.cell(row=row, column=14, value=p.url)

    # Sütun genişlikleri
    col_widths = [12, 50, 20, 25, 15, 18, 12, 10, 15, 10, 25, 16, 20, 50]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"

    # --- Yorumlar Sayfası (isteğe bağlı) ---
    if reviews:
        ws2 = wb.create_sheet("Yorumlar")
        r_headers = ["ID", "Ürün ID", "Yazar", "Puan", "Başlık", "Yorum", "Tarih", "Beğeni", "Beğenmeme", "Doğrulanmış"]
        for col, h in enumerate(r_headers, 1):
            cell = ws2.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill

        for row, r in enumerate(reviews, 2):
            ws2.cell(row=row, column=1, value=r.id)
            ws2.cell(row=row, column=2, value=r.product_id)
            ws2.cell(row=row, column=3, value=r.author)
            ws2.cell(row=row, column=4, value=r.rating)
            ws2.cell(row=row, column=5, value=r.title)
            ws2.cell(row=row, column=6, value=r.comment)
            ws2.cell(row=row, column=7, value=r.date)
            ws2.cell(row=row, column=8, value=r.likes)
            ws2.cell(row=row, column=9, value=r.dislikes)
            ws2.cell(row=row, column=10, value="Evet" if r.is_verified else "Hayır")

    Path(filepath).parent.mkdir(parents=True, exist_ok=True)
    wb.save(filepath)
    logger.info(f"Excel kaydedildi: {filepath}")
